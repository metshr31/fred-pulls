import os
import datetime as dt
import re
import pandas as pd
import pdfplumber
from openpyxl import load_workbook, Workbook
from typing import List

DOWNLOAD_FOLDER = os.getenv("STB_LOG_DIR", os.getcwd())

def datestamp() -> str:
    return dt.date.today().strftime("%Y-%m-%d")

# -------------------------
# Parsing functions (PDFs)
# -------------------------
def parse_ns_carloads(pdf_path: str) -> pd.DataFrame:
    records = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                df = pd.DataFrame(table)
                if df.shape[0] > 2 and "This Yr" in " ".join(map(str, df.iloc[1])):
                    df.columns = ["Commodity","CW_ThisYr","CW_LastYr","CW_Delta",
                                  "QTD_ThisYr","QTD_LastYr","QTD_Delta",
                                  "YTD_ThisYr","YTD_LastYr","YTD_Delta"]
                    df = df.drop([0,1]).reset_index(drop=True)
                    records.extend(df.to_dict(orient="records"))
    return pd.DataFrame(records)

def parse_bnsf_carloads(pdf_path: str) -> pd.DataFrame:
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                if "|" in line and "%" in line:
                    parts = [p.strip() for p in line.split("|") if p.strip()]
                    if len(parts) != 3:
                        continue
                    left, right, pct = parts
                    m = re.match(r"([A-Za-z/&' \-0-9]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)", left)
                    if not m:
                        continue
                    category, y2025_w, y2025_q, y2025_y = m.groups()
                    nums2024 = right.split()
                    if len(nums2024) < 3:
                        continue
                    y2024_w, y2024_q, y2024_y = nums2024[:3]
                    pcts = [tok for tok in pct.split() if "%" in tok]
                    if len(pcts) < 3:
                        continue
                    rows.append({
                        "Category": category.strip(),
                        "2025_Week": y2025_w, "2025_QTD": y2025_q, "2025_YTD": y2025_y,
                        "2024_Week": y2024_w, "2024_QTD": y2024_q, "2024_YTD": y2024_y,
                        "Pct_Week": pcts[0], "Pct_QTD": pcts[1], "Pct_YTD": pcts[2]
                    })
    return pd.DataFrame(rows)

def parse_csx_aar(pdf_path: str) -> pd.DataFrame:
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join([page.extract_text() or "" for page in pdf.pages])
    pattern = re.compile(r"([A-Za-z&\(\)\/ \-]+)\s+([\d,]+)\s+([\d,]+)\s+\(?([\-0-9\.]+%?)\)?\s+([\d,]+)\s+([\d,]+)\s+\(?([\-0-9\.]+%?)\)?\s+([\d,]+)\s+([\d,]+)\s+\(?([\-0-9\.]+%?)\)?")
    for m in pattern.finditer(text):
        rows.append({
            "Category": m.group(1).strip(),
            "2025_Week": m.group(2), "2024_Week": m.group(3), "Pct_Week": m.group(4),
            "2025_QTD": m.group(5), "2024_QTD": m.group(6), "Pct_QTD": m.group(7),
            "2025_YTD": m.group(8), "2024_YTD": m.group(9), "Pct_YTD": m.group(10)
        })
    return pd.DataFrame(rows)

# -------------------------
# Merge functions
# -------------------------
def merge_to_master_excel(fetched: List[str]) -> str:
    out_file = os.path.join(DOWNLOAD_FOLDER, f"rail_service_master_{datestamp()}.xlsx")
    with pd.ExcelWriter(out_file, engine="xlsxwriter") as writer:
        for f in fetched:
            if f.endswith(".pdf"):
                if "NS_Carloads" in f:
                    parse_ns_carloads(f).to_excel(writer, sheet_name="NS_Carloads_pdf", index=False)
                elif "BNSF_Carloads" in f:
                    parse_bnsf_carloads(f).to_excel(writer, sheet_name="BNSF_Carloads_pdf", index=False)
                elif "CSX_AAR" in f:
                    parse_csx_aar(f).to_excel(writer, sheet_name="CSX_Carloads_pdf", index=False)
            elif f.endswith(".xlsx"):
                xls = pd.ExcelFile(f)
                for sheet in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet, header=None)
                    sheet_name = (os.path.basename(f).replace(".xlsx","") + "_" + sheet)[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
    print(f"📊 Master Excel written: {out_file}")
    return out_file

def merge_excels_with_formatting(fetched: List[str]) -> str:
    out_file = os.path.join(DOWNLOAD_FOLDER, f"rail_service_excels_merged_{datestamp()}.xlsx")
    merged_wb = Workbook()
    merged_wb.remove(merged_wb.active)
    added_any = False

    for f in fetched:
        if not f.endswith(".xlsx"):
            continue
        try:
            wb = load_workbook(f)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                new_sheet_name = (os.path.splitext(os.path.basename(f))[0] + "_" + sheet_name)[:31]
                new_ws = merged_wb.create_sheet(title=new_sheet_name)

                for row in ws.iter_rows():
                    for cell in row:
                        new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell._style = cell._style

                for col, dim in ws.column_dimensions.items():
                    new_ws.column_dimensions[col].width = dim.width
                for row, dim in ws.row_dimensions.items():
                    new_ws.row_dimensions[row].height = dim.height
                for merged_range in ws.merged_cells.ranges:
                    new_ws.merge_cells(str(merged_range))
                added_any = True

            print(f"✅ Merged {f}")
        except Exception as e:
            print(f"⚠️ Could not process {f}: {e}")

    if not added_any:
        merged_wb.create_sheet(title="Empty")

    merged_wb.save(out_file)
    print(f"📊 Excel-only merged workbook written: {out_file}")
    return out_file

# -------------------------
# Main
# -------------------------
def main():
    fetched = [os.path.join(DOWNLOAD_FOLDER, f) for f in os.listdir(DOWNLOAD_FOLDER)]
    merge_to_master_excel(fetched)
    merge_excels_with_formatting(fetched)

if __name__ == "__main__":
    main()
